""" import """
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from tslearn.clustering import TimeSeriesKMeans 
import xlsxwriter 
from openpyxl import load_workbook
import os

""" parameter """
# get file 
file = 'Annual_timeseries.xlsx'
# define sites (Viertel2, WU, Stadion, Neubau)
sites = 4
# define types (Elec, Heat, Cold, Solar, Heat Pump Efficiency)
types = 5
# dimensions of matrix
number = sites * types
hours = 168
weeks = 52
h_w = hours * weeks


""" subfunctions """


# read in excel file
def read_excel_file(file):
    return pd.read_excel(file)


# reshape column vector to shape (168, 52)
def reshape_timeserie(cv):
    cv = cv[0:h_w]
    euclidean_norm = np.linalg.norm(cv)
    cv = cv / euclidean_norm
    return np.reshape(cv,(hours, weeks))


# plot time serie, input is matrix to plot and number of columns
def plot_timeseries(matrix, c):
    length=range(hours)
    plt.figure(figsize = (16,9))
    for i in range(c):
        plt.plot(length, matrix[:, i])


# sort one type of time series
def sort_one_type(matrix):
    global sites
    matrix = matrix[0:h_w, :]
    for i in range(sites):
        vector = matrix[:, i]
        euclidean_norm = np.linalg.norm(vector)
        matrix[:, i] = vector/euclidean_norm
    return np.reshape(matrix, (hours, sites*weeks), order='F')


# generates demand sheet "Demand_Final.xlsx"
def generate_demand_sheet(fn1, fn2, fn3):
    elec = pd.read_excel(fn1)
    heat = pd.read_excel(fn2)
    cold = pd.read_excel(fn3)
    dataframe = pd.concat((elec,heat,cold), axis=1)
    dataframe.to_excel("Demand_Final.xlsx", sheet_name='Demand')
    return


# write t value to excel file for better input to urbs 
def write_t(name,sheet):
    wb = load_workbook(name)
    ws = wb[sheet]
    ws['A1'].value = 't'
    wb.save(name)
    wb.close()


# generates solar sheet "Solar_Final.xlsx"   
def generate_solar_sheet(fn1):
    solar = pd.read_excel(fn1)
    solar.to_excel("Solar_Final.xlsx", sheet_name='SupIm')
    write_t("Solar_Final.xlsx", 'SupIm')
    return 


# generates solar sheet "Solar_Final.xlsx" 
def generate_e_sheet(fn1):
    solar = pd.read_excel(fn1)
    solar.to_excel("TimeVarEff_Final.xlsx", sheet_name='TimeVarEff')
    write_t("TimeVarEff_Final.xlsx", 'TimeVarEff')
    return 


def plot_ts_and_cluster(result, source, string):
    # result: matrix of clustered ts
    # source 52 weeks input data
    global g
    length = range(hours)
    plt.figure(figsize = (16,9))
    plt.xlabel('Hours [h]', fontsize=14)
    plt.ylabel('Euclidean norm', fontsize=14)
    plt.title('Clustering ' +string +' with '+str(g)+' cluster', fontsize=18)
    for i in range(weeks):
        plt.plot(length, source[:, i], linewidth=0.2, color='black')
    """ define linewidth with weight of cluster"""
    for j in range(g):
        linewidth = weight[j]*5/np.max(weight)
        plt.plot(length, result[:, j], marker='x', linestyle='dashed', linewidth=linewidth, markersize=3)
    plt.savefig('0_'+string+'.png')
    return


# start "main" function for clustering algorithm
file = read_excel_file(file)
list_names = list(file)
cv = file.values
length = range(hours)
maximum = np.zeros((1, sites*types))
array = np.array(file)


for i in range(sites*types):
    maximum[0, i] = np.linalg.norm(cv[0:hours*weeks, i])
del[array, i]
maximum = np.reshape(maximum, (types, sites))


# build input matrix
elec = sort_one_type(cv[:, 0:1*sites])
heat = sort_one_type(cv[:, sites:2*sites])
cold = sort_one_type(cv[:, 2*sites:3*sites])
solar = sort_one_type(cv[:, 3*sites:4*sites])
e = sort_one_type(cv[:, 4*sites:5*sites])

""" 
matrix for kmeans need (n x p) with n samples (52 weeks x 4 sites = 208)
and p features (168h * 5 types = 840)
"""

matrix_all = np.concatenate((elec, heat, cold, solar, e), axis = 0)
matrix_to_kmeans = np.transpose(matrix_all)

current_dir = os.path.dirname(__file__)

# edit number of clusters to represent yearly timeseries
for g in range(1):
    g = g+3
    try:
        os.chdir('Results_clusters='+str(g))
    except:
        os.mkdir('Results_clusters='+str(g))
        os.chdir('Results_clusters='+str(g))
        
    #print(os.getcwd())
    
    km = TimeSeriesKMeans(n_clusters = g).fit(matrix_to_kmeans)
    centroids = np.matrix(km.cluster_centers_).transpose()
    labels = np.matrix(km.labels_)
    weight = np.zeros((g, 1))
    for i in range(g):
        weight[i] = np.sum(labels == i)/sites
    string = ['Elec', 'Heat', 'Cold', 'Solar', 'Efficiency']
    for i in range(types):
        workbook = xlsxwriter.Workbook(string[i]+'_results.xlsx')
        worksheet = workbook.add_worksheet()
        matrix_to_write = centroids[i*hours:(i+1)*hours, :]
        matrix_to_write = np.reshape(matrix_to_write, (g*hours, 1), order='F')
        for j in range(g*hours):
            for q in range(sites):
                worksheet.write(j+1, q, matrix_to_write[j, 0]*maximum[i, q])
        for u in range(sites):
            worksheet.write(0, u, list_names[i*sites+u])
        workbook.close()
    workbook = xlsxwriter.Workbook('Weight_Final.xlsx')
    worksheet = workbook.add_worksheet('Weight')
    for r in range(g):
        for j in range(hours):
            worksheet.write(r*hours+j + 1, 0, weight[r])
    worksheet.write(0,0,'Weight')    
    workbook.close()
       
    generate_demand_sheet('Elec_results.xlsx', 'Heat_results.xlsx',\
                      'Cold_results.xlsx')

    write_t('Demand_Final.xlsx', 'Demand')
    generate_solar_sheet('Solar_results.xlsx')
    generate_e_sheet('Efficiency_results.xlsx')

    for z in range(types):
        plot_ts_and_cluster(centroids[z*hours:(z+1)*hours, :], \
                                  matrix_all[z*hours:(z+1)*hours, :], \
                                  (string[z]))
    os.chdir(current_dir)